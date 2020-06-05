import cv2
from global_var import personGroupId
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string
import time
import pickle



	
face_cascade = cv2.CascadeClassifier('cascades/data/haarcascade_frontalface_alt2.xml')
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("trainner.yml") 

labels = {"person_name":1}

a = int(input("Enter the Number of people"))

currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports_ise.xlsx")
sheet = wb.active
#sheet = ws['ISE17']


def getDateColumn():
	for i in range(1, len(sheet[1]) + 1):
		col = i
		row=1
		
		if sheet.cell(row,col).value == currentDate:
			return col
	
		
		
connect = connect = sqlite3.connect("Face_DataBase")
c = connect.cursor()

attend = [0 for i in range(100)]	

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'dataset')
labels = {"person_name":1}
img_dir=os.path.join(currentDir,"test")
with open("labels.pickle","rb") as f:
	og_labels = pickle.load(f)
	labels = {v:k for  k,v in og_labels.items()}






cap=cv2.VideoCapture(0)
k=0
while(True):
	count=0
	ret, frame=cap.read()
	gray =cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
	faces =face_cascade.detectMultiScale(gray, scaleFactor= 1.05,minNeighbors=4 )
	cv2.imshow('frame',frame)
	
	for (x, y, w, h) in faces:
		#print (x,y,w,h)
		count=count+1
		roi_gray = gray[y:y+h,x:x+w] #(cord1-height, cord2-height)
		roi_color = gray[y:y+h,x:x+w]
		id_,conf =recognizer.predict(roi_color)
		
		if conf>=0 and conf<=50:
			#print(id_)
			confidence=100-conf
			
			font = cv2.FONT_HERSHEY_SIMPLEX
			name = labels[id_]+"  "+ str(int(confidence))+"%"
			color = (255 ,255,255)
			stroke =2
			cv2.putText(frame,name ,(x,y),font,1,color,stroke,cv2.LINE_AA)
			
			
			color = (255 ,0, 0)
			stroke=2
			end_cord_x= x+w
			end_cord_y= y+h
			cv2.rectangle(frame, (x,y), (end_cord_x,end_cord_y),color ,stroke)
			#print (labels[id_])
			
			#for i in range(0,count):
					
			#face = faces[i]
			name = labels[id_]	
				#print (face)
				
			personId = name
				#print(personId)
			c.execute("SELECT * FROM Students WHERE personid = ?", (personId,))
			row = c.fetchone()
				#print (row)
			if ( attend[int(row[0])] ==1):
				continue
			else:
				attend[int(row[0])] += 1
				print (row[1] + " recognized")
				print(confidence)
				k=k+1
			
			
		else:
			
			continue
				
	cv2.imshow('frame',frame)
	if cv2.waitKey(20) & 0xFF ==ord('q'):
		break
	elif k==a:
		break		
	
print ("face detected count=",count)				

for row in range(2, len(sheet["A"])+1):
	column=column_index_from_string("A")
	
	rn = sheet.cell(row,column).value
	#print (rn)
	if rn is not None:
		
		rn = sheet.cell(row,column).value
		
		if attend[int(rn)] != 0 :
			print("Marked Present")
			col = (getDateColumn())
			cols=get_column_letter(col)+str(row)
			#print(cols)
			sheet[cols] = 1
		else:
			col = (getDateColumn())
			cols=get_column_letter(col)+str(row)
			#print(cols)
			sheet[cols] = 0
wb.save(filename = "reports_ise.xlsx")
	 	

cap.release()
cv2.destroyAllWindows()
			
