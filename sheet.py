from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string
import time
import os
import sqlite3

#database connection
conn = sqlite3.connect('Face_DataBase')
c = conn.cursor()

#get current date
currentDate = time.strftime("%d_%m_%y")

#create a workbook and add a worksheet
if(os.path.exists('./reports_ise.xlsx')):
    wb = load_workbook(filename = "reports_ise.xlsx")
    sheet = wb['ISE17']
    # sheet[ord() + '1']
    for col_index in range(3, 100):
    	col = get_column_letter(col_index)
    	print(sheet.cell(row=1,column=col_index).value)
    	if (sheet.cell(row=1,column=col_index).value == None):
    		#col2 = get_column_letter(col_index+1)
    		# print sheet.cell('%s%s'% (col2, 1)).value
    		if sheet.cell(1,col_index - 1 ).value != currentDate:
    			sheet['%s%s' % (col,1)] = currentDate
    		break

    #saving the file
    wb.save(filename = "reports_ise.xlsx")
    	
else:
    wb = Workbook()
    dest_filename = 'reports_ise.xlsx'
    c.execute("SELECT * FROM Students ORDER BY Roll ASC")
    
    #creating worksheet and giving names to column
    ws1 = wb.active
    ws1.title = "ISE17"
    ws1.append(('Roll Number', 'Name', currentDate))
    ws1.append(('', '', ''))

    #entering students information from database
    while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            ws1.append((a[2], a[1]))

    #saving the file
    wb.save(filename = dest_filename)   
